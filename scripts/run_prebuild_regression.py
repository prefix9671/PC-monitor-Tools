from __future__ import annotations

import json
import subprocess
import sys
import time
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

ARTIFACT_DIR = REPO_ROOT / ".artifacts" / "prebuild-regression"
SUMMARY_PATH = ARTIFACT_DIR / "prebuild-regression-summary.json"
PLAYWRIGHT_URL = "http://127.0.0.1:8506"
PLAYWRIGHT_PORT = 8506
PYTHON_EXE = Path(sys.executable)

BUG_OPERATION_LOG = REPO_ROOT / "bug" / "operation_0319_north side grab.log"
INSPECTOR_TIME_FILTER_FIXTURE_LOG = REPO_ROOT / "tests" / "fixtures" / "inspector_time_filter_range_regression.log"
SPI_TACT_TIME_FIXTURE_CSV = REPO_ROOT / "tests" / "fixtures" / "spi_tact_time_regression.csv"
SPI_PROCESS_RESOURCE_FIXTURE_LOG = REPO_ROOT / "tests" / "fixtures" / "spi_process_resource_regression.log"
BUG_RESOURCE_CSV = REPO_ROOT / "bug" / "20260410-메모리 대시보드 버그" / "resource_20260410.csv"
BUG_PROCESS_CSV = REPO_ROOT / "bug" / "20260410-메모리 대시보드 버그" / "process_20260410.csv"
AOI_EXPORT_PATH = ARTIFACT_DIR / "inspection_export_prebuild.xlsx"


class StepFailure(RuntimeError):
    pass


def _print_step_header(name: str, description: str, failure_condition: str) -> None:
    print(f"[STEP] {name}")
    print(f"CHECK: {description}")
    print(f"FAILS IF: {failure_condition}")


def _print_step_footer(name: str, ok: bool) -> None:
    print(f"[{'PASS' if ok else 'FAIL'}] {name}")


def _write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def _sanitize_console_text(text: str) -> str:
    encoding = sys.stdout.encoding or "utf-8"
    return text.encode(encoding, errors="replace").decode(encoding, errors="replace")


def _run_command_step(
    summary_steps: list[dict[str, object]],
    name: str,
    description: str,
    failure_condition: str,
    command: list[str],
) -> subprocess.CompletedProcess[str]:
    _print_step_header(name, description, failure_condition)
    print(f"COMMAND: {' '.join(command)}")

    completed = subprocess.run(
        command,
        cwd=REPO_ROOT,
        check=False,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )

    stdout_path = ARTIFACT_DIR / f"{name}.stdout.log"
    stderr_path = ARTIFACT_DIR / f"{name}.stderr.log"
    _write_text(stdout_path, completed.stdout)
    _write_text(stderr_path, completed.stderr)

    print("STDOUT:")
    print(_sanitize_console_text(completed.stdout.rstrip()) or "<empty>")
    if completed.stderr.strip():
        print("STDERR:")
        print(_sanitize_console_text(completed.stderr.rstrip()))

    step_record = {
        "name": name,
        "description": description,
        "failure_condition": failure_condition,
        "command": command,
        "returncode": completed.returncode,
        "stdout_path": str(stdout_path),
        "stderr_path": str(stderr_path),
        "ok": completed.returncode == 0,
    }
    summary_steps.append(step_record)

    if completed.returncode != 0:
        _print_step_footer(name, ok=False)
        raise StepFailure(f"{name} failed with exit code {completed.returncode}.")

    _print_step_footer(name, ok=True)
    return completed


def _run_python_check_step(
    summary_steps: list[dict[str, object]],
    name: str,
    description: str,
    failure_condition: str,
    checker,
) -> None:
    _print_step_header(name, description, failure_condition)
    stdout_path = ARTIFACT_DIR / f"{name}.stdout.log"
    stderr_path = ARTIFACT_DIR / f"{name}.stderr.log"
    try:
        stdout_lines = checker()
        stdout_text = "\n".join(stdout_lines)
        _write_text(stdout_path, stdout_text)
        _write_text(stderr_path, "")
        print("STDOUT:")
        print(_sanitize_console_text(stdout_text.rstrip()) or "<empty>")
        summary_steps.append(
            {
                "name": name,
                "description": description,
                "failure_condition": failure_condition,
                "stdout_path": str(stdout_path),
                "stderr_path": str(stderr_path),
                "ok": True,
            }
        )
        _print_step_footer(name, ok=True)
    except Exception as exc:
        error_text = str(exc)
        _write_text(stdout_path, "")
        _write_text(stderr_path, error_text)
        print("STDERR:")
        print(_sanitize_console_text(error_text))
        summary_steps.append(
            {
                "name": name,
                "description": description,
                "failure_condition": failure_condition,
                "stdout_path": str(stdout_path),
                "stderr_path": str(stderr_path),
                "ok": False,
                "error": error_text,
            }
        )
        _print_step_footer(name, ok=False)
        raise StepFailure(f"{name} failed: {error_text}") from exc


def _wait_for_streamlit_ready(url: str, timeout_seconds: int = 60) -> None:
    deadline = time.time() + timeout_seconds
    last_error = None
    while time.time() < deadline:
        try:
            with urllib.request.urlopen(url, timeout=2) as response:
                if response.status == 200:
                    return
        except Exception as exc:
            last_error = exc
        time.sleep(1)
    raise StepFailure(f"Streamlit server did not become ready: {last_error}")


def _run_playwright_step(summary_steps: list[dict[str, object]]) -> None:
    name = "headless-playwright-regression"
    description = (
        "Headless Streamlit + Playwright regression verifies dashboard rendering plus "
        "AOI/SPI upload, filter, and download flows."
    )
    failure_condition = (
        "Streamlit server fails to start, Playwright cannot connect, any UI step returns non-zero, "
        "AOI time filter does not shrink the NO range, or AOI/SPI downloads are missing."
    )
    _print_step_header(name, description, failure_condition)

    streamlit_stdout_path = ARTIFACT_DIR / "streamlit.stdout.log"
    streamlit_stderr_path = ARTIFACT_DIR / "streamlit.stderr.log"
    streamlit_stdout = streamlit_stdout_path.open("w", encoding="utf-8")
    streamlit_stderr = streamlit_stderr_path.open("w", encoding="utf-8")
    streamlit_process = None

    try:
        streamlit_process = subprocess.Popen(
            [
                str(PYTHON_EXE),
                "-m",
                "streamlit",
                "run",
                "app.py",
                "--server.headless",
                "true",
                "--server.port",
                str(PLAYWRIGHT_PORT),
            ],
            cwd=REPO_ROOT,
            stdout=streamlit_stdout,
            stderr=streamlit_stderr,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        _wait_for_streamlit_ready(PLAYWRIGHT_URL)

        completed = subprocess.run(
            [
                "node",
                str(REPO_ROOT / "scripts" / "verify_playwright_prebuild_regression.js"),
                PLAYWRIGHT_URL,
                str(BUG_RESOURCE_CSV),
                str(BUG_PROCESS_CSV),
                str(BUG_OPERATION_LOG),
                str(SPI_TACT_TIME_FIXTURE_CSV),
                str(SPI_PROCESS_RESOURCE_FIXTURE_LOG),
            ],
            cwd=REPO_ROOT,
            check=False,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )

        stdout_path = ARTIFACT_DIR / f"{name}.stdout.log"
        stderr_path = ARTIFACT_DIR / f"{name}.stderr.log"
        _write_text(stdout_path, completed.stdout)
        _write_text(stderr_path, completed.stderr)

        print("STDOUT:")
        print(_sanitize_console_text(completed.stdout.rstrip()) or "<empty>")
        if completed.stderr.strip():
            print("STDERR:")
            print(_sanitize_console_text(completed.stderr.rstrip()))

        summary_steps.append(
            {
                "name": name,
                "description": description,
                "failure_condition": failure_condition,
                "command": [
                    "node",
                    "scripts/verify_playwright_prebuild_regression.js",
                    PLAYWRIGHT_URL,
                    str(BUG_RESOURCE_CSV),
                    str(BUG_PROCESS_CSV),
                    str(BUG_OPERATION_LOG),
                    str(SPI_TACT_TIME_FIXTURE_CSV),
                    str(SPI_PROCESS_RESOURCE_FIXTURE_LOG),
                ],
                "returncode": completed.returncode,
                "stdout_path": str(stdout_path),
                "stderr_path": str(stderr_path),
                "streamlit_stdout_path": str(streamlit_stdout_path),
                "streamlit_stderr_path": str(streamlit_stderr_path),
                "ok": completed.returncode == 0,
            }
        )

        if completed.returncode != 0:
            _print_step_footer(name, ok=False)
            raise StepFailure(f"{name} failed with exit code {completed.returncode}.")

        _print_step_footer(name, ok=True)
    finally:
        if streamlit_process is not None and streamlit_process.poll() is None:
            streamlit_process.terminate()
            try:
                streamlit_process.wait(timeout=15)
            except subprocess.TimeoutExpired:
                streamlit_process.kill()
                streamlit_process.wait(timeout=15)
        streamlit_stdout.close()
        streamlit_stderr.close()


def _verify_inputs_exist() -> list[str]:
    required_paths = [
        BUG_OPERATION_LOG,
        INSPECTOR_TIME_FILTER_FIXTURE_LOG,
        SPI_TACT_TIME_FIXTURE_CSV,
        SPI_PROCESS_RESOURCE_FIXTURE_LOG,
        BUG_RESOURCE_CSV,
        BUG_PROCESS_CSV,
    ]
    missing = [str(path) for path in required_paths if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Missing regression input files: {missing}")
    return [f"inputs_ready={len(required_paths)}", *(str(path) for path in required_paths)]


def _verify_inspector_time_filter_fixture() -> list[str]:
    from inspector_logs.core import (
        build_inspection_records,
        filter_inspection_records_by_time_range,
        load_inspector_log_data,
        select_inspection_records,
        summarize_inspection_records,
    )

    start_time = "2026-05-13 15:00:00"
    end_time = "2026-05-13 16:44:59"
    expected_rows = 3
    expected_no_range = (2, 4)

    inspector_df = load_inspector_log_data(str(INSPECTOR_TIME_FILTER_FIXTURE_LOG))
    inspection_records = build_inspection_records(inspector_df)
    filtered_records = filter_inspection_records_by_time_range(
        inspection_records,
        start_time=start_time,
        end_time=end_time,
    )
    summary = summarize_inspection_records(filtered_records)

    if summary["rows"] != expected_rows:
        raise AssertionError(f"Expected {expected_rows} filtered rows, got {summary['rows']}.")
    if summary["no_range"] != expected_no_range:
        raise AssertionError(f"Expected NO range {expected_no_range}, got {summary['no_range']}.")

    selected_records = select_inspection_records(
        filtered_records,
        start_no=expected_no_range[0],
        end_no=expected_no_range[1],
    )
    if len(selected_records) != expected_rows:
        raise AssertionError(f"Expected selected rows {expected_rows}, got {len(selected_records)}.")

    return [
        f"log={INSPECTOR_TIME_FILTER_FIXTURE_LOG}",
        f"time_filter={start_time} -> {end_time}",
        f"parsed_events={len(inspector_df)}",
        f"inspection_records={len(inspection_records)}",
        f"filtered_rows={summary['rows']}",
        f"filtered_no_range={summary['no_range']}",
        f"selected_rows={len(selected_records)}",
    ]


def _verify_aoi_export_workbook() -> list[str]:
    if not AOI_EXPORT_PATH.exists():
        raise FileNotFoundError(f"Expected workbook was not created: {AOI_EXPORT_PATH}")

    workbook = load_workbook(AOI_EXPORT_PATH, read_only=True)
    sheet_names = workbook.sheetnames
    if "Inspection_Results" not in sheet_names or "Inspection_12h_Samples" not in sheet_names:
        raise AssertionError(f"Unexpected workbook sheets: {sheet_names}")

    sample_sheet = workbook["Inspection_12h_Samples"]
    return [
        f"sheet_names={sheet_names}",
        f"A1={sample_sheet['A1'].value}",
        f"B1={sample_sheet['B1'].value}",
        f"A5={sample_sheet['A5'].value}",
    ]


def main() -> int:
    ARTIFACT_DIR.mkdir(parents=True, exist_ok=True)
    summary_steps: list[dict[str, object]] = []

    try:
        _run_python_check_step(
            summary_steps,
            name="preflight-inputs",
            description="Repo-local regression inputs exist for system CSVs, AOI log, and SPI logs.",
            failure_condition="Any required regression fixture under bug/ or tests/fixtures/ is missing.",
            checker=_verify_inputs_exist,
        )

        _run_command_step(
            summary_steps,
            name="unit-tests",
            description="Core AOI export, inspection panel, time filtering, memory dashboard, and packaging tests pass.",
            failure_condition="The unittest command exits non-zero.",
            command=[
                str(PYTHON_EXE),
                "-m",
                "unittest",
                "tests.test_inspector_logs",
                "tests.test_aoi_cli",
                "tests.test_inspection_export_panel",
                "tests.test_time_filtering",
                "tests.test_memory_dashboard",
                "tests.test_packaging_layout",
            ],
        )

        _run_command_step(
            summary_steps,
            name="aoi-cli-summary",
            description="AOI CLI summary parses the repo-local bug log.",
            failure_condition="aoi_cli summary exits non-zero or cannot parse the bug log.",
            command=[
                str(PYTHON_EXE),
                "aoi_cli.py",
                "summary",
                "--path",
                str(BUG_OPERATION_LOG),
            ],
        )

        _run_python_check_step(
            summary_steps,
            name="aoi-inspector-time-filter-fixture-regression",
            description="Minimal inspector fixture time filter keeps all matching inspection rows and the full NO range.",
            failure_condition="Filtered row count is not 3, NO range is not 2 -> 4, or selected rows collapse to 1.",
            checker=_verify_inspector_time_filter_fixture,
        )

        _run_command_step(
            summary_steps,
            name="aoi-cli-export",
            description="AOI CLI export creates the regression workbook with the repo-local bug log.",
            failure_condition="aoi_cli export exits non-zero or the workbook is not generated.",
            command=[
                str(PYTHON_EXE),
                "aoi_cli.py",
                "export",
                "--path",
                str(BUG_OPERATION_LOG),
                "--out",
                str(AOI_EXPORT_PATH),
            ],
        )

        _run_python_check_step(
            summary_steps,
            name="aoi-export-workbook-check",
            description="Generated workbook contains Inspection_Results and Inspection_12h_Samples sheets.",
            failure_condition="Workbook is missing or sheet layout does not match the expected export contract.",
            checker=_verify_aoi_export_workbook,
        )

        _run_command_step(
            summary_steps,
            name="dashboard-smoke",
            description="Headless dashboard smoke test passes with repo-local system CSV fixtures.",
            failure_condition="verify_dashboards.py exits non-zero or any dashboard crashes on the fixture logs.",
            command=[
                str(PYTHON_EXE),
                "verify_dashboards.py",
                "--files",
                str(BUG_RESOURCE_CSV),
                str(BUG_PROCESS_CSV),
            ],
        )

        _run_command_step(
            summary_steps,
            name="docs-sync",
            description="Active docs stay in sync with the changed code paths.",
            failure_condition="scripts/verify_docs_sync.py exits non-zero.",
            command=[str(PYTHON_EXE), "scripts/verify_docs_sync.py"],
        )

        _run_command_step(
            summary_steps,
            name="mkdocs-build-check",
            description="MkDocs site build succeeds before packaging.",
            failure_condition="python -m mkdocs build exits non-zero.",
            command=[str(PYTHON_EXE), "-m", "mkdocs", "build"],
        )

        _run_playwright_step(summary_steps)

        summary = {
            "ok": True,
            "checked_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
            "artifact_dir": str(ARTIFACT_DIR),
            "steps": summary_steps,
        }
        SUMMARY_PATH.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[DONE] Pre-build regression summary: {SUMMARY_PATH}")
        return 0
    except Exception as exc:
        summary = {
            "ok": False,
            "checked_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
            "artifact_dir": str(ARTIFACT_DIR),
            "error": str(exc),
            "steps": summary_steps,
        }
        SUMMARY_PATH.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[ERROR] Pre-build regression failed: {exc}")
        print(f"[ERROR] Summary: {SUMMARY_PATH}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
