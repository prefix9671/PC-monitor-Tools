import os
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from aoi_cli import main
from excel_exporter import generate_inspection_excel


class TestAoiCli(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.temp_path = Path(self.temp_dir.name)

        self.aoi_log_path = self.temp_path / "inspection_session.log"
        self.aoi_log_path.write_text(
            "\n".join(
                [
                    "20260318_174405 | debug    | (22376) |Model Open : SAMPLE_MODEL_A",
                    "20260318_174406 | debug    | (105760) |Memory|Working Set Memory Size | 47487856 KB",
                    "20260318_174407 | debug    | (22376) |InspTime|Frame : 1.19 sec|Total : 38.13 sec / 32 frame",
                    "20260318_174411 | debug    | (105760) |Memory|Working Set Memory Size | 50000000 KB",
                    "20260318_174412 | debug    | (22376) |InspTime|Frame : 1.50 sec|Total : 40.00 sec / 33 frame",
                ]
            ),
            encoding="utf-8",
        )

        self.resource_csv_path = self.temp_path / "resource_20260318.csv"
        pd.DataFrame(
            {
                "Timestamp": ["2026-03-18 17:44:06", "2026-03-18 17:44:10"],
                "CPU_Avg(%)": [10.0, 12.0],
                "Mem_Used(GB)": [28.5, 29.0],
                "Mem_Usage_Avg(%)": [50.0, 51.0],
            }
        ).to_csv(self.resource_csv_path, index=False, encoding="utf-8-sig")

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_export_command_generates_xlsx_with_numbered_rows(self):
        output_path = self.temp_path / "inspection_export.xlsx"
        test_args = [
            "aoi_cli.py",
            "export",
            "--path",
            str(self.aoi_log_path),
            "--system-path",
            str(self.resource_csv_path),
            "--start-no",
            "1",
            "--end-no",
            "2",
            "--out",
            str(output_path),
        ]

        with patch.object(sys, "argv", test_args):
            exit_code = main()

        self.assertEqual(0, exit_code)
        self.assertTrue(output_path.exists())

        export_df = pd.read_excel(output_path)
        workbook = load_workbook(output_path)
        self.assertEqual(
            ["NO", "Frame", "Total", "메모리 (시스템)"],
            export_df.columns.tolist(),
        )
        self.assertEqual([1, 2], export_df["NO"].tolist())
        self.assertAlmostEqual(28.5, float(export_df.iloc[0]["메모리 (시스템)"]), places=2)
        self.assertAlmostEqual(29.0, float(export_df.iloc[1]["메모리 (시스템)"]), places=2)
        self.assertEqual(["Inspection_Results", "Inspection_12h_Samples"], workbook.sheetnames)
        sample_sheet = workbook["Inspection_12h_Samples"]
        self.assertEqual("적용 시작 시각", sample_sheet["A1"].value)
        self.assertEqual("적용 종료 시각", sample_sheet["A2"].value)
        self.assertEqual("종료 시점 지정", sample_sheet["A3"].value)
        self.assertTrue(str(sample_sheet["A5"].value).startswith("+0h 샘플"))

    def test_export_command_can_optionally_include_inspector_memory(self):
        output_path = self.temp_path / "inspection_export_with_inspector.xlsx"
        test_args = [
            "aoi_cli.py",
            "export",
            "--path",
            str(self.aoi_log_path),
            "--system-path",
            str(self.resource_csv_path),
            "--include-inspector-memory",
            "--out",
            str(output_path),
        ]

        with patch.object(sys, "argv", test_args):
            exit_code = main()

        self.assertEqual(0, exit_code)
        export_df = pd.read_excel(output_path)
        workbook = load_workbook(output_path)
        self.assertEqual(
            ["NO", "Frame", "Total", "메모리 (시스템)", "메모리 (인스펙터)"],
            export_df.columns.tolist(),
        )
        self.assertAlmostEqual(47487856 / (1024.0 * 1024.0), float(export_df.iloc[0]["메모리 (인스펙터)"]), places=3)
        sample_sheet = workbook["Inspection_12h_Samples"]
        self.assertEqual(
            ["Timestamp", "NO", "Frame", "Total", "메모리 (시스템)", "메모리 (인스펙터)"],
            [sample_sheet["A6"].value, sample_sheet["B6"].value, sample_sheet["C6"].value, sample_sheet["D6"].value, sample_sheet["E6"].value, sample_sheet["F6"].value],
        )

    def test_generate_inspection_excel_uses_sample_records_for_additional_sheet(self):
        full_records = pd.DataFrame(
            {
                "Timestamp": pd.to_datetime(
                    [
                        "2026-03-18 00:00:05",
                        "2026-03-18 12:00:05",
                    ]
                ),
                "SourceFile": ["sample.log", "sample.log"],
                "Inspection_No": [1, 2],
                "Inspector_Model_Name": ["SAMPLE_MODEL_A", "SAMPLE_MODEL_A"],
                "Inspector_Frame_Sec": [1.0, 1.2],
                "Inspector_Total_Sec": [10.0, 12.0],
                "Inspector_Total_Frames": [10, 10],
                "Inspector_WorkingSet_KB": [1000.0, 1200.0],
                "Inspector_WorkingSet_MB": [1.0, 1.2],
                "Inspector_WorkingSet_GB": [0.001, 0.0012],
                "System_Memory_Used_GB": [20.0, 22.0],
                "System_Memory_Timestamp": pd.to_datetime(
                    [
                        "2026-03-18 00:00:00",
                        "2026-03-18 12:00:00",
                    ]
                ),
            }
        )
        selected_records = full_records.iloc[[1]].reset_index(drop=True)
        output_path = self.temp_path / "inspection_export_sample_scope.xlsx"

        output_path.write_bytes(
            generate_inspection_excel(
                selected_records,
                sample_records=full_records,
                sample_start_time=pd.Timestamp("2026-03-18 00:00:00"),
                sample_end_time=pd.Timestamp("2026-03-18 12:30:00"),
            )
        )

        workbook = load_workbook(output_path)
        sample_sheet = workbook["Inspection_12h_Samples"]
        self.assertEqual(pd.Timestamp("2026-03-18 00:00:00").to_pydatetime(), sample_sheet["B1"].value)
        self.assertEqual(1, sample_sheet["B7"].value)


if __name__ == "__main__":
    unittest.main()
