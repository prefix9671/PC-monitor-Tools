import os
import sys
import unittest
from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from dashboards.inspection_export import (
    COLOR_PRESET_MAP,
    INSPECTION_EXPORT_END_NO_KEY,
    INSPECTION_EXPORT_RANGE_SCOPE_KEY,
    INSPECTION_EXPORT_START_NO_KEY,
    _build_preview_chart,
    _build_inspection_xlsx_download_payload,
    _build_inspection_xlsx_download_key,
    _format_time_filter_caption,
    _hex_to_rgba,
    _make_inspection_export_scope,
    _resolve_inspection_no_range_state,
)


class TestInspectionExportPanel(unittest.TestCase):
    def setUp(self):
        self.preview_df = pd.DataFrame(
            {
                "NO": [1, 2],
                "Frame": [1.19, 1.50],
                "Total": [38.13, 40.00],
                "메모리 (시스템)": [28.5, 29.0],
                "메모리 (인스펙터)": [45.288, 47.684],
            }
        )

    def test_hex_to_rgba(self):
        self.assertEqual("rgba(58, 134, 255, 0.28)", _hex_to_rgba("#3A86FF", 0.28))

    def test_named_palette_count_stays_within_requested_range(self):
        self.assertLessEqual(len(COLOR_PRESET_MAP), 36)

    def test_build_preview_chart_line_mode(self):
        fig = _build_preview_chart(
            preview_df=self.preview_df,
            chart_type="선 + 마커",
            selected_metrics=["Frame", "Total", "메모리 (시스템)", "메모리 (인스펙터)"],
            metric_colors={
                "Frame": COLOR_PRESET_MAP["플럼 바이올렛"],
                "Total": COLOR_PRESET_MAP["선셋 오렌지"],
                "메모리 (시스템)": COLOR_PRESET_MAP["코발트 블루"],
                "메모리 (인스펙터)": COLOR_PRESET_MAP["에메랄드"],
            },
            opacity=0.85,
        )

        self.assertEqual(4, len(fig.data))
        self.assertTrue(all(trace.type == "scatter" for trace in fig.data))
        self.assertEqual("검사 시간 (sec)", fig.layout.yaxis.title.text)
        self.assertEqual("메모리 (GB)", fig.layout.yaxis2.title.text)
        self.assertEqual("NO", fig.layout.xaxis.title.text)

    def test_build_preview_chart_bar_mode(self):
        fig = _build_preview_chart(
            preview_df=self.preview_df,
            chart_type="막대",
            selected_metrics=["Frame", "메모리 (시스템)", "메모리 (인스펙터)"],
            metric_colors={
                "Frame": COLOR_PRESET_MAP["플럼 바이올렛"],
                "메모리 (시스템)": COLOR_PRESET_MAP["코발트 블루"],
                "메모리 (인스펙터)": COLOR_PRESET_MAP["에메랄드"],
            },
            opacity=0.55,
        )

        self.assertEqual(3, len(fig.data))
        self.assertTrue(all(trace.type == "bar" for trace in fig.data))

    def test_format_time_filter_caption(self):
        caption = _format_time_filter_caption(
            pd.Timestamp("2026-03-18 00:00:00"),
            pd.Timestamp("2026-03-18 12:00:00"),
        )

        self.assertIn("현재 시간 필터 기준", caption)
        self.assertIn("2026-03-18 00:00:00", caption)
        self.assertIn("2026-03-18 12:00:00", caption)

    def test_no_range_state_resets_to_filtered_full_range_when_filter_scope_changes(self):
        state = {
            INSPECTION_EXPORT_RANGE_SCOPE_KEY: "old-filter",
            INSPECTION_EXPORT_START_NO_KEY: 2,
            INSPECTION_EXPORT_END_NO_KEY: 2,
        }

        start_no, end_no = _resolve_inspection_no_range_state(
            state,
            min_no=2,
            max_no=4,
            row_count=3,
            filter_start_time=pd.Timestamp("2026-05-13 15:00:00"),
            filter_end_time=pd.Timestamp("2026-05-13 16:44:59"),
        )

        self.assertEqual(2, start_no)
        self.assertEqual(4, end_no)
        self.assertEqual(4, state[INSPECTION_EXPORT_END_NO_KEY])

    def test_no_range_state_preserves_manual_selection_inside_same_filter_scope(self):
        scope = _make_inspection_export_scope(
            min_no=2,
            max_no=4,
            row_count=3,
            filter_start_time=pd.Timestamp("2026-05-13 15:00:00"),
            filter_end_time=pd.Timestamp("2026-05-13 16:44:59"),
        )
        state = {
            INSPECTION_EXPORT_RANGE_SCOPE_KEY: scope,
            INSPECTION_EXPORT_START_NO_KEY: 3,
            INSPECTION_EXPORT_END_NO_KEY: 4,
        }

        start_no, end_no = _resolve_inspection_no_range_state(
            state,
            min_no=2,
            max_no=4,
            row_count=3,
            filter_start_time=pd.Timestamp("2026-05-13 15:00:00"),
            filter_end_time=pd.Timestamp("2026-05-13 16:44:59"),
        )

        self.assertEqual(3, start_no)
        self.assertEqual(4, end_no)

    def test_xlsx_payload_includes_inspector_memory_when_checked(self):
        records = pd.DataFrame(
            {
                "Timestamp": pd.to_datetime(["2026-07-06 14:00:01", "2026-07-06 14:00:31"]),
                "Inspection_No": [1, 2],
                "Inspector_Frame_Sec": [0.45, 0.47],
                "Inspector_Total_Sec": [23.4, 23.5],
                "System_Memory_Used_GB": [30.1, 30.2],
                "Inspector_WorkingSet_GB": [2.901, 2.908],
            }
        )

        payload = _build_inspection_xlsx_download_payload(
            selected_records=records,
            selected_model="SPI",
            start_no=1,
            end_no=2,
            include_inspector_memory=True,
            sample_records=records,
            sample_start_time=pd.Timestamp("2026-07-06 14:00:00"),
            sample_end_time=pd.Timestamp("2026-07-06 14:01:00"),
            generated_at=datetime(2026, 7, 6, 14, 2, 3),
        )

        workbook = load_workbook(BytesIO(payload["data"]))
        result_headers = [cell.value for cell in workbook["Inspection_Results"][1]]
        sample_headers = [workbook["Inspection_12h_Samples"][f"{column}6"].value for column in "ABCDEF"]

        self.assertEqual(
            "Inspection_Results_SPI_NO0001-0002_20260706_140203.xlsx",
            payload["file_name"],
        )
        self.assertIn("메모리 (인스펙터)", result_headers)
        self.assertEqual(
            ["Timestamp", "NO", "Frame", "Total", "메모리 (시스템)", "메모리 (인스펙터)"],
            sample_headers,
        )

    def test_xlsx_payload_omits_inspector_memory_when_unchecked(self):
        records = pd.DataFrame(
            {
                "Timestamp": pd.to_datetime(["2026-07-06 14:00:01"]),
                "Inspection_No": [1],
                "Inspector_Frame_Sec": [0.45],
                "Inspector_Total_Sec": [23.4],
                "System_Memory_Used_GB": [30.1],
                "Inspector_WorkingSet_GB": [2.901],
            }
        )

        payload = _build_inspection_xlsx_download_payload(
            selected_records=records,
            selected_model="SPI",
            start_no=1,
            end_no=1,
            include_inspector_memory=False,
            sample_records=records,
            generated_at=datetime(2026, 7, 6, 14, 2, 3),
        )

        workbook = load_workbook(BytesIO(payload["data"]))
        result_headers = [cell.value for cell in workbook["Inspection_Results"][1]]

        self.assertNotIn("메모리 (인스펙터)", result_headers)

    def test_xlsx_download_key_changes_with_inspector_memory_option(self):
        unchecked_key = _build_inspection_xlsx_download_key(
            start_no=1,
            end_no=2,
            selected_row_count=2,
            filtered_row_count=10,
            include_inspector_memory=False,
        )
        checked_key = _build_inspection_xlsx_download_key(
            start_no=1,
            end_no=2,
            selected_row_count=2,
            filtered_row_count=10,
            include_inspector_memory=True,
        )

        self.assertNotEqual(unchecked_key, checked_key)
        self.assertIn("without-inspector-memory", unchecked_key)
        self.assertIn("with-inspector-memory", checked_key)


if __name__ == "__main__":
    unittest.main()
