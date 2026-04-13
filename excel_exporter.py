import pandas as pd
import io
import re
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from inspector_logs.core import (
    build_inspection_sample_sections,
    format_inspection_export_dataframe,
    resolve_inspection_sample_time_window,
)

def parse_top5_string(data_str):
    """문자열 형태의 Top5 데이터를 리스트로 변환 (예: 'proc1:100MB | proc2:50MB')"""
    if not data_str or str(data_str).lower() in ["no_active_io", "nan", "none"]:
        return []
    
    items = []
    for item in [x.strip() for x in str(data_str).split('|') if x.strip()]:
        if ':' in item:
            name, val = item.split(':', 1)
            items.append((name.strip(), val.strip()))
    return items

def generate_excel(df, selected_cols):
    """
    selected_cols 및 프로세스 데이터를 포함하여 엑셀 파일을 생성합니다.
    """
    output = io.BytesIO()
    
    # 1. 시계열 지표 데이터 준비
    export_df = df[['Timestamp'] + selected_cols].copy()
    
    # 2. 프로세스 데이터 추가 (데이터가 존재하는 경우)
    # 메모리 프로세스
    if 'Top5_Memory_MB' in df.columns:
        for i in range(5):
            export_df[f'Top_Mem_Proc_{i+1}'] = df['Top5_Memory_MB'].apply(
                lambda x: parse_top5_string(x)[i][0] if len(parse_top5_string(x)) > i else ""
            )
            export_df[f'Top_Mem_Val_{i+1}'] = df['Top5_Memory_MB'].apply(
                lambda x: parse_top5_string(x)[i][1] if len(parse_top5_string(x)) > i else ""
            )

    # 디스크 프로세스
    if 'Top5_Disk_IO_Global(MB/s)' in df.columns:
        for i in range(5):
            export_df[f'Top_Disk_Proc_{i+1}'] = df['Top5_Disk_IO_Global(MB/s)'].apply(
                lambda x: parse_top5_string(x)[i][0] if len(parse_top5_string(x)) > i else ""
            )
            export_df[f'Top_Disk_Val_{i+1}'] = df['Top5_Disk_IO_Global(MB/s)'].apply(
                lambda x: parse_top5_string(x)[i][1] if len(parse_top5_string(x)) > i else ""
            )
    
    # 컬럼 순서 조정 (Timestamp를 '값'으로 표시하거나 유지)
    export_df.rename(columns={'Timestamp': '시간(Timestamp)'}, inplace=True)
    
    def sanitize_str(val):
        if isinstance(val, str):
            # Remove control characters except \n, \t, \r
            return ''.join(c for c in val if ord(c) >= 32 or ord(c) in (9, 10, 13))
        return val

    # Sanitize dataframe columns and string data
    export_df.columns = [sanitize_str(c) for c in export_df.columns]
    for col in export_df.select_dtypes(include=['object']):
        export_df[col] = export_df[col].apply(sanitize_str)

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        export_df.to_excel(writer, index=False, sheet_name='System_Resource_Report')
        
    return output.getvalue()


def _apply_column_widths(writer, sheet_name, export_df):
    worksheet = writer.book[sheet_name]
    for column_index, column_name in enumerate(export_df.columns, start=1):
        column_values = [column_name, *export_df[column_name].astype(str).tolist()]
        max_length = min(max(len(str(value)) for value in column_values) + 2, 28)
        worksheet.column_dimensions[get_column_letter(column_index)].width = max_length


def _apply_column_widths_from_cells(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = None
        for cell in column_cells:
            column_letter = cell.column_letter
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        if column_letter is not None:
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 32)


def _write_inspection_sample_sheet(
    writer,
    sample_records,
    include_inspector_memory=False,
    sample_start_time=None,
    sample_end_time=None,
    end_time_user_specified=False,
):
    worksheet = writer.book.create_sheet(title="Inspection_12h_Samples")
    effective_start, effective_end = resolve_inspection_sample_time_window(
        sample_records,
        start_time=sample_start_time,
        end_time=sample_end_time,
    )

    worksheet.append(["적용 시작 시각", effective_start])
    worksheet.append(["적용 종료 시각", effective_end])
    worksheet.append(["종료 시점 지정", "사용자 지정" if end_time_user_specified else "자동/기본 범위"])
    worksheet.append([])

    for section in build_inspection_sample_sections(
        sample_records,
        start_time=effective_start,
        end_time=effective_end,
        include_inspector_memory=include_inspector_memory,
    )["sections"]:
        anchor_timestamp = section["anchor_timestamp"]
        title = f"+{section['anchor_hours']}h 샘플 ({anchor_timestamp.strftime('%Y-%m-%d %H:%M:%S')})"
        worksheet.append([title])
        worksheet[worksheet.max_row][0].font = Font(bold=True)

        if section["status"] == "rows":
            dataframe = section["dataframe"]
            worksheet.append(list(dataframe.columns))
            for cell in worksheet[worksheet.max_row]:
                cell.font = Font(bold=True)
            for row in dataframe.itertuples(index=False, name=None):
                worksheet.append(list(row))
        else:
            worksheet.append(["안내", section["message"]])
            for cell in worksheet[worksheet.max_row]:
                cell.font = Font(bold=True)

        worksheet.append([])

    _apply_column_widths_from_cells(worksheet)


def generate_inspection_excel(
    inspection_records,
    include_inspector_memory=False,
    sample_records=None,
    sample_start_time=None,
    sample_end_time=None,
    end_time_user_specified=False,
):
    """
    검사 결과 레코드를 XLSX로 내보냅니다.
    """
    output = io.BytesIO()
    export_df = format_inspection_export_dataframe(
        inspection_records,
        include_inspector_memory=include_inspector_memory,
    )
    effective_sample_records = inspection_records if sample_records is None else sample_records

    with pd.ExcelWriter(output, engine="openpyxl", datetime_format="YYYY-MM-DD HH:MM:SS") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Inspection_Results")
        _apply_column_widths(writer, "Inspection_Results", export_df)
        _write_inspection_sample_sheet(
            writer,
            effective_sample_records,
            include_inspector_memory=include_inspector_memory,
            sample_start_time=sample_start_time,
            sample_end_time=sample_end_time,
            end_time_user_specified=end_time_user_specified,
        )

    return output.getvalue()

